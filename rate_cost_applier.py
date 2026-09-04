from __future__ import annotations

import re
import shutil
from pathlib import Path

from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from cost_mappings import CostMappings, load_cost_mappings
from paths import OUTPUT_DIR, PROCESSING_DIR, RATE_AGREEMENT_INPUT_DIR
from rate_card_layouts import (
    build_rate_lookup,
    describe_detected_layout,
    list_warehouse_columns_from_file,
    resolve_warehouse_price_col,
    warehouse_label_for_column,
)

AGREEMENT_INPUT_DIR = RATE_AGREEMENT_INPUT_DIR
RATE_CARD_PROCESSING_DIR = PROCESSING_DIR
RATE_CARD_SHEET = "Rate card"
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
VALUE_HEADERS = {"p/unit", "p / unit", "flat"}
HANDLING_SERVICE = "HANDLING"
STORAGE_SERVICE = "STORAGE"
DEFAULT_SERVICES = [HANDLING_SERVICE, STORAGE_SERVICE]
NIGHT_SHIFT_SUFFIX = " [night shift]"
NIGHT_QUANTITY_AVAILABLE_PATTERN = re.compile(
    r"quantity/night\s+is\s+available",
    re.IGNORECASE,
)
NIGHT_SHIFT_COMMENT_PATTERN = re.compile(
    r"comment:\s*night\s+shift",
    re.IGNORECASE,
)
MINIMUM_CHARGE_HEADER_PATTERN = re.compile(r"^min$", re.IGNORECASE)
MINIMUM_CHARGE_LOOKUP_KEYS = (
    "minimum charge für HANDLING OUT",
    "minimum charge for HANDLING OUT",
    "minimum charge for handling out (up to 0,0313m³)",
)
THRESHOLD_PATTERN = re.compile(r"(?P<op><=|>=|≤|≥|<|>)\s*(?P<value>[\d.,]+)")
VOLUME_RATE_HINTS = ("volume/cbm", "volume/cbn", "0.002")
SMALL_VALUE_THRESHOLD = 0.01


def list_agreement_files() -> list[Path]:
    """Return available rate agreement xlsx files from the input folder."""
    if not AGREEMENT_INPUT_DIR.exists():
        raise FileNotFoundError(f"Input folder not found: {AGREEMENT_INPUT_DIR}")

    files = sorted(AGREEMENT_INPUT_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in: {AGREEMENT_INPUT_DIR}")

    return files


def list_rate_card_files() -> list[Path]:
    """Return processed rate card xlsx files from the processing folder."""
    if not RATE_CARD_PROCESSING_DIR.exists():
        raise FileNotFoundError(f"Processing folder not found: {RATE_CARD_PROCESSING_DIR}")

    files = sorted(RATE_CARD_PROCESSING_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in: {RATE_CARD_PROCESSING_DIR}")

    return files


def choose_file(files: list[Path], label: str) -> Path:
    """Prompt the user to choose a file from a numbered list."""
    print(f"\nAvailable {label} files:")
    for index, file_path in enumerate(files, start=1):
        print(f"  {index}. {file_path.name}")

    while True:
        choice = input("Enter file number: ").strip()
        if not choice.isdigit():
            print("Please enter a valid number.")
            continue

        selected_index = int(choice)
        if 1 <= selected_index <= len(files):
            return files[selected_index - 1]

        print(f"Please enter a number between 1 and {len(files)}.")


def normalize_key(value: str) -> str:
    """Normalize lookup keys for case-insensitive matching."""
    text = re.sub(r"\s+", " ", value.strip().lower())
    text = text.rstrip(":").strip()
    text = re.sub(r"[-–—]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_number(value) -> float | None:
    """Parse numeric values, including European decimal commas."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def extract_lookup_keys(header_texts: list[str]) -> list[str]:
    """
    Build candidate lookup keys from agreement header rows.

    Handles:
    - Category names with parentheses, e.g. 'Handling Fee (Backflush)' -> 'Backflush'
    - Apply-if descriptions that contain the actual rate card cost name
    """
    keys: list[str] = []

    for text in header_texts:
        text = str(text).strip()
        if not text:
            continue

        keys.append(text)

        apply_if_match = re.search(r"apply\s*if[:\s]+(.+)", text, re.IGNORECASE)
        if apply_if_match:
            apply_if_text = apply_if_match.group(1).strip()
            keys.append(apply_if_text)
            text = apply_if_text

        for match in re.finditer(r"\(([^)]+)\)", text):
            keys.append(match.group(1).strip())

        if " - " in text:
            for part in text.split(" - "):
                cleaned = part.split("(")[0].strip()
                if cleaned:
                    keys.append(cleaned)

    seen: set[str] = set()
    unique_keys: list[str] = []
    for key in keys:
        normalized = normalize_key(key)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_keys.append(key)

    return unique_keys


def parse_threshold_tier_from_text(text: str) -> str | None:
    """Return the tier for a single threshold expression."""
    match = THRESHOLD_PATTERN.search(str(text))
    if not match:
        return None

    operator = match.group("op")
    if operator in ("<", "<=", "≤"):
        return "lower"
    if operator in (">", ">=", "≥"):
        return "upper"

    return None


def parse_threshold_tier(header_texts: list[str]) -> str | None:
    """
    Detect tier type from threshold text in header rows.

    Returns 'lower' for </<= thresholds and 'upper' for >/>= thresholds.
    """
    tiers: list[str] = []

    for text in header_texts:
        matches = list(THRESHOLD_PATTERN.finditer(str(text)))
        if len(matches) == 1:
            tier = parse_threshold_tier_from_text(matches[0].group(0))
            if tier:
                tiers.append(tier)

    if not tiers:
        return None

    unique_tiers = set(tiers)
    if len(unique_tiers) == 1:
        return tiers[0]

    return None


def get_category_block_key(ws: Worksheet, col: int, category_row: int) -> str:
    """Return a stable key for the row-1 cost category block that owns this column."""
    value = get_merged_cell_value(ws, category_row, col)
    if value is None:
        return f"col_{col}"

    return normalize_key(str(value).strip())


def get_column_threshold_tier(
    ws: Worksheet,
    col: int,
    category_row: int,
    header_row: int,
) -> str | None:
    """
    Detect the threshold tier for one column.

    Scans header rows from bottom to top and prefers values stored directly in the column.
    """
    for row in range(header_row - 1, category_row - 1, -1):
        direct_value = ws.cell(row, col).value
        candidates = []
        if direct_value is not None and str(direct_value).strip():
            candidates.append(str(direct_value).strip())

        merged_value = get_merged_cell_value(ws, row, col)
        if merged_value is not None and str(merged_value).strip():
            merged_text = str(merged_value).strip()
            if merged_text not in candidates:
                candidates.append(merged_text)

        for text in candidates:
            tier = parse_threshold_tier([text])
            if tier:
                return tier

    return None


def get_pair_group_key(lookup_keys: list[str], header_texts: list[str]) -> str:
    """Build a stable group key for related tiered cost column pairs."""
    for key in lookup_keys:
        if THRESHOLD_PATTERN.search(key):
            continue

        normalized = normalize_key(key)
        if normalized:
            return normalized

    for text in header_texts:
        if THRESHOLD_PATTERN.search(text):
            continue

        normalized = normalize_key(text)
        if normalized:
            return normalized

    return normalize_key(header_texts[0]) if header_texts else ""


def is_value_header(ws: Worksheet, header_row: int, col: int) -> bool:
    """Return True when a column header is a cost value field."""
    header = str(ws.cell(header_row, col).value or "").strip().lower()
    return header in VALUE_HEADERS


def collect_following_value_columns(
    ws: Worksheet,
    header_row: int,
    start_col: int,
) -> list[int]:
    """Collect consecutive value columns that follow a currency column."""
    value_columns: list[int] = []
    col = start_col

    while col <= ws.max_column and is_value_header(ws, header_row, col):
        value_columns.append(col)
        col += 1

    return value_columns


def group_has_volume_tiers(pairs: list[dict[str, object]], indices: list[int]) -> bool:
    """Return True when a cost block uses volume threshold columns."""
    for index in indices:
        pair = pairs[index]
        header_text = normalize_key(" | ".join(str(text) for text in pair["header_texts"]))
        if pair.get("threshold_tier"):
            return True
        if any(hint in header_text for hint in VOLUME_RATE_HINTS):
            return True

    return False


def mark_tiered_groups(pairs: list[dict[str, object]]) -> None:
    """Mark column pairs that belong to a lower/upper threshold tier group."""
    groups: dict[str, list[int]] = defaultdict(list)

    for index, pair in enumerate(pairs):
        groups[str(pair["category_block_key"])].append(index)

    for indices in groups.values():
        tiers = {
            str(pairs[index]["threshold_tier"])
            for index in indices
            if pairs[index].get("threshold_tier")
        }
        is_tiered = "lower" in tiers and "upper" in tiers

        if not is_tiered and len(indices) >= 2 and group_has_volume_tiers(pairs, indices):
            is_tiered = True

        if not is_tiered:
            continue

        sorted_indices = sorted(indices, key=lambda index: int(pairs[index]["value_col"]))
        for rank, index in enumerate(sorted_indices):
            pair = pairs[index]
            pair["is_tiered"] = True

            if pair.get("threshold_tier") is None and len(sorted_indices) == 2:
                pair["threshold_tier"] = "lower" if rank == 0 else "upper"
            elif pair.get("threshold_tier") is None and rank == 0:
                pair["threshold_tier"] = "lower"
            elif pair.get("threshold_tier") is None:
                pair["threshold_tier"] = "upper"


def extract_lookup_key(category_name: str | None) -> str:
    """Return the primary lookup key for backward compatibility."""
    keys = extract_lookup_keys([str(category_name or "").strip()])
    return keys[-1] if keys else ""


def register_lookup_key(lookup: dict[str, float], key: str, value: float) -> None:
    """Register a lookup key and its normalized form."""
    if not key:
        return

    lookup[key] = value
    lookup[normalize_key(key)] = value


def _lookup_partial_tokens(text: str) -> list[str]:
    return [
        token
        for token in re.split(r"[\s\-/()+]+", normalize_key(text))
        if len(token) >= 4
    ]


def header_is_night_shift(header_texts: list[str]) -> bool:
    """Return True only for columns with Quantity/Night and Comment: Night Shift."""
    blob = normalize_key(" | ".join(str(text) for text in header_texts if str(text).strip()))
    if not blob:
        return False

    return bool(
        NIGHT_QUANTITY_AVAILABLE_PATTERN.search(blob)
        and NIGHT_SHIFT_COMMENT_PATTERN.search(blob)
    )


def header_is_minimum_charge(header_texts: list[str]) -> bool:
    """Return True when the agreement column is a minimum-charge cell."""
    for text in header_texts:
        if MINIMUM_CHARGE_HEADER_PATTERN.fullmatch(str(text).strip()):
            return True

    return False


def with_night_shift_suffix(name: str) -> str:
    return f"{name}{NIGHT_SHIFT_SUFFIX}"


def lookup_rate_card_value(
    rate_card_name: str,
    rate_lookup: dict[str, float],
    prefer_night_shift: bool = False,
) -> tuple[float | None, str | None]:
    """Look up a rate card value by exact, normalized, or partial key match."""
    candidate_names: list[str] = []
    if prefer_night_shift:
        candidate_names.append(with_night_shift_suffix(rate_card_name))
    candidate_names.append(rate_card_name)

    for candidate in candidate_names:
        for key in (candidate, normalize_key(candidate)):
            if key in rate_lookup:
                return rate_lookup[key], candidate

    norm_name = normalize_key(candidate_names[0])
    if not norm_name:
        return None, None

    night_suffix = normalize_key(NIGHT_SHIFT_SUFFIX)

    for rate_key, value in rate_lookup.items():
        if not isinstance(rate_key, str):
            continue

        norm_rate_key = normalize_key(rate_key)
        if not prefer_night_shift and norm_rate_key.endswith(night_suffix):
            continue

        if norm_name == norm_rate_key:
            return value, rate_key

        shorter, longer = sorted((norm_name, norm_rate_key), key=len)
        if len(shorter) >= 12 and (shorter in longer or longer in shorter):
            return value, rate_key

        tokens = _lookup_partial_tokens(candidate_names[0])
        if len(tokens) >= 2 and all(token in norm_rate_key for token in tokens):
            return value, rate_key

        if len(tokens) == 1 and len(tokens[0]) >= 8 and tokens[0] in norm_rate_key:
            return value, rate_key

    if not prefer_night_shift:
        return None, None

    return None, None


def lookup_minimum_charge_value(
    rate_lookup: dict[str, float],
    prefer_night_shift: bool = False,
) -> tuple[float | None, str | None]:
    """Look up the DENV minimum handling-out charge."""
    for base_name in MINIMUM_CHARGE_LOOKUP_KEYS:
        value, matched_key = lookup_rate_card_value(
            base_name,
            rate_lookup,
            prefer_night_shift=prefer_night_shift,
        )
        if value is not None:
            return value, matched_key

    return None, None


def is_small_placeholder_value(value) -> bool:
    """Return True for tiny agreement values that should stay unchanged."""
    numeric = parse_number(value)
    if numeric is None:
        return False

    return abs(numeric) < SMALL_VALUE_THRESHOLD


def find_rate_value(
    lookup_keys: list[str],
    rate_lookup: dict[str, float],
    header_texts: list[str] | None = None,
    cost_mappings: CostMappings | None = None,
    prefer_night_shift: bool = False,
    prefer_minimum_charge: bool = False,
) -> tuple[float | None, str | None, str | None]:
    """
    Find a rate card value.

    Priority:
    1. Cost mappings txt
    2. Existing agreement header/name matching
    """
    header_texts = header_texts or []
    night_suffix = normalize_key(NIGHT_SHIFT_SUFFIX)

    if prefer_minimum_charge:
        value, matched_key = lookup_minimum_charge_value(
            rate_lookup,
            prefer_night_shift=prefer_night_shift,
        )
        if value is not None:
            return value, matched_key, "minimum"

    if cost_mappings and cost_mappings.entries:
        mapped_names = cost_mappings.agreement_candidates(header_texts, lookup_keys)
        for rate_card_name in mapped_names:
            value, matched_key = lookup_rate_card_value(
                rate_card_name,
                rate_lookup,
                prefer_night_shift=prefer_night_shift,
            )
            if value is not None:
                return value, matched_key, "mapping"

    for candidate in lookup_keys:
        for key in (
            with_night_shift_suffix(candidate) if prefer_night_shift else candidate,
            normalize_key(with_night_shift_suffix(candidate) if prefer_night_shift else candidate),
        ):
            if key in rate_lookup:
                return rate_lookup[key], candidate, "name"

    for candidate in lookup_keys:
        if prefer_night_shift:
            continue

        norm_candidate = normalize_key(candidate)
        if not norm_candidate:
            continue

        for rate_key, value in rate_lookup.items():
            if not isinstance(rate_key, str):
                continue

            norm_rate_key = normalize_key(rate_key)
            if norm_rate_key.endswith(night_suffix):
                continue

            if norm_candidate in norm_rate_key or norm_rate_key in norm_candidate:
                return value, rate_key, "name"

            tokens = _lookup_partial_tokens(candidate)
            if len(tokens) >= 2 and all(token in norm_rate_key for token in tokens):
                return value, rate_key, "name"

            if len(tokens) == 1 and len(tokens[0]) >= 8 and tokens[0] in norm_rate_key:
                return value, rate_key, "name"

    if prefer_night_shift:
        for candidate in lookup_keys:
            norm_candidate = normalize_key(candidate)
            if not norm_candidate:
                continue

            for rate_key, value in rate_lookup.items():
                if not isinstance(rate_key, str):
                    continue

                norm_rate_key = normalize_key(rate_key)
                if not norm_rate_key.endswith(night_suffix):
                    continue

                base_rate_key = norm_rate_key[: -len(normalize_key(NIGHT_SHIFT_SUFFIX))].strip()
                if norm_candidate in base_rate_key or base_rate_key in norm_candidate:
                    return value, rate_key, "name"

                tokens = _lookup_partial_tokens(candidate)
                if len(tokens) >= 2 and all(token in base_rate_key for token in tokens):
                    return value, rate_key, "name"

    return None, None, None


def get_merged_cell_value(ws: Worksheet, row: int, col: int):
    """Return the value of a cell, resolving merged ranges to the top-left cell."""
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return ws.cell(row, col).value


def find_header_row(ws: Worksheet) -> int:
    """Find the row that contains the Service and Lane column headers."""
    for row in range(1, min(20, ws.max_row + 1)):
        values = [
            str(ws.cell(row, col).value or "").strip().lower()
            for col in range(1, ws.max_column + 1)
        ]
        if "service" in values and any("lane" in value for value in values):
            return row

    return 4


def find_column_by_header(ws: Worksheet, header_row: int, header_name: str) -> int | None:
    """Find a column index by its header label."""
    target = header_name.strip().lower()
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value and str(value).strip().lower() == target:
            return col
    return None


def get_column_header_texts(
    ws: Worksheet,
    currency_col: int,
    value_col: int,
    start_row: int,
    end_row: int,
) -> list[str]:
    """Collect header text for a currency/value pair from all header rows above the data header."""
    texts: list[str] = []
    seen: set[str] = set()

    for col in (currency_col, value_col):
        for row in range(start_row, end_row):
            direct_value = ws.cell(row, col).value
            values = []
            if direct_value is not None and str(direct_value).strip():
                values.append(str(direct_value).strip())

            merged_value = get_merged_cell_value(ws, row, col)
            if merged_value is not None and str(merged_value).strip():
                merged_text = str(merged_value).strip()
                if merged_text not in values:
                    values.append(merged_text)

            for text in values:
                normalized = normalize_key(text)
                if normalized in seen:
                    continue

                seen.add(normalized)
                texts.append(text)

    return texts


def build_cost_column_pairs(
    ws: Worksheet,
    category_row: int,
    header_row: int,
) -> list[dict[str, int | str | list[str]]]:
    """Map currency/value column pairs to their cost category lookup keys."""
    pairs: list[dict[str, int | str | list[str]]] = []
    col = 1

    while col <= ws.max_column:
        header_value = str(ws.cell(header_row, col).value or "").strip().lower()
        if header_value != "currency":
            col += 1
            continue

        currency_col = col
        value_columns = collect_following_value_columns(ws, header_row, col + 1)
        if not value_columns:
            col += 1
            continue

        for value_col in value_columns:
            header_texts = get_column_header_texts(
                ws,
                currency_col=currency_col,
                value_col=value_col,
                start_row=category_row,
                end_row=header_row,
            )
            lookup_keys = extract_lookup_keys(header_texts)
            threshold_tier = get_column_threshold_tier(ws, value_col, category_row, header_row)
            if threshold_tier is None:
                threshold_tier = get_column_threshold_tier(ws, currency_col, category_row, header_row)
            if threshold_tier is None:
                threshold_tier = parse_threshold_tier(header_texts)

            category_block_key = get_category_block_key(ws, value_col, category_row)
            if category_block_key.startswith("col_"):
                category_block_key = get_category_block_key(ws, currency_col, category_row)

            group_key = get_pair_group_key(lookup_keys, header_texts)

            pairs.append(
                {
                    "currency_col": currency_col,
                    "value_col": value_col,
                    "category": " | ".join(header_texts),
                    "header_texts": header_texts,
                    "lookup_keys": lookup_keys,
                    "threshold_tier": threshold_tier,
                    "group_key": group_key,
                    "category_block_key": category_block_key,
                    "is_tiered": False,
                }
            )

        col = value_columns[-1] + 1

    return pairs



def service_matches_target(service_value, target_services: set[str]) -> bool:
    """Return True when a row service matches one of the requested services."""
    normalized = str(service_value).strip().upper()
    if normalized in target_services:
        return True

    parts = re.split(r"[/+|\\]", normalized)
    return any(part.strip() in target_services for part in parts)


def apply_costs_to_service_rows(
    ws: Worksheet,
    rate_lookup: dict[str, float],
    services: list[str],
    cost_mappings: CostMappings | None = None,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Apply rate card costs to selected service rows and highlight updated cells."""
    header_row = find_header_row(ws)
    category_row = 1
    service_col = find_column_by_header(ws, header_row, "Service")
    if service_col is None:
        raise ValueError("Could not find a 'Service' column in the rate agreement sheet.")

    cost_pairs = build_cost_column_pairs(ws, category_row, header_row)
    if not cost_pairs:
        raise ValueError("Could not find any currency/value cost column pairs.")

    mark_tiered_groups(cost_pairs)

    target_services = {service.strip().upper() for service in services}
    updates: list[dict[str, object]] = []
    unmatched: list[dict[str, object]] = []

    for row in range(header_row + 1, ws.max_row + 1):
        service_value = ws.cell(row, service_col).value
        if service_value is None:
            continue

        if not service_matches_target(service_value, target_services):
            continue

        for pair in cost_pairs:
            currency_col = int(pair["currency_col"])
            value_col = int(pair["value_col"])
            value_cell = ws.cell(row, value_col)
            currency_cell = ws.cell(row, currency_col)
            current_value = value_cell.value

            if current_value is None or str(current_value).strip() == "":
                continue

            is_tiered = bool(pair.get("is_tiered"))
            threshold_tier = pair.get("threshold_tier")

            if is_tiered and threshold_tier == "lower":
                currency_cell.fill = YELLOW_FILL
                value_cell.fill = YELLOW_FILL

                updates.append(
                    {
                        "row": row,
                        "column": value_col,
                        "service": str(service_value).strip(),
                        "category": pair["category"],
                        "lookup_key": None,
                        "old_value": current_value,
                        "new_value": current_value,
                        "action": "highlighted_yellow",
                    }
                )
                continue

            if is_small_placeholder_value(current_value):
                currency_cell.fill = YELLOW_FILL
                value_cell.fill = YELLOW_FILL

                updates.append(
                    {
                        "row": row,
                        "column": value_col,
                        "service": str(service_value).strip(),
                        "category": pair["category"],
                        "lookup_key": None,
                        "old_value": current_value,
                        "new_value": current_value,
                        "action": "highlighted_yellow",
                    }
                )
                continue

            if is_tiered and threshold_tier == "upper":
                lookup_keys = [
                    key
                    for key in pair["lookup_keys"]
                    if not THRESHOLD_PATTERN.search(str(key))
                ]
            else:
                lookup_keys = list(pair["lookup_keys"])

            header_texts = list(pair.get("header_texts", []))
            prefer_night_shift = header_is_night_shift(header_texts)
            prefer_minimum_charge = header_is_minimum_charge(header_texts)
            new_value, matched_key, match_source = find_rate_value(
                lookup_keys,
                rate_lookup,
                header_texts=header_texts,
                cost_mappings=cost_mappings,
                prefer_night_shift=prefer_night_shift,
                prefer_minimum_charge=prefer_minimum_charge,
            )
            if new_value is None:
                unmatched.append(
                    {
                        "row": row,
                        "column": value_col,
                        "service": str(service_value).strip(),
                        "category": pair["category"],
                        "lookup_keys": lookup_keys,
                    }
                )
                continue

            value_cell.value = new_value
            value_cell.fill = GREEN_FILL

            updates.append(
                {
                    "row": row,
                    "column": value_col,
                    "service": str(service_value).strip(),
                    "category": pair["category"],
                    "lookup_key": matched_key,
                    "match_source": match_source,
                    "old_value": current_value,
                    "new_value": new_value,
                    "action": "updated_green",
                }
            )

    return updates, unmatched


def apply_rate_card_to_agreement(
    agreement_path: Path,
    rate_card_path: Path,
    sheet_name: str = RATE_CARD_SHEET,
    services: list[str] | None = None,
    mappings_path: Path | None = None,
    warehouse: str | int | None = None,
    *,
    interactive: bool = True,
) -> tuple[Path, list[dict[str, object]], list[dict[str, object]]]:
    """
    Apply rate card costs to a rate agreement file and save to the output folder.

    The original workbook formatting is preserved. Updated value cells are highlighted green.
    """
    services = services or DEFAULT_SERVICES
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    cost_mappings = load_cost_mappings(mappings_path)
    print(f"Loaded {len(cost_mappings.entries)} cost mappings from txt")
    print(f"Detected rate card layout: {describe_detected_layout(rate_card_path)}")

    warehouses = list_warehouse_columns_from_file(rate_card_path)
    warehouse_price_col = resolve_warehouse_price_col(
        warehouses,
        warehouse,
        interactive=interactive,
    )
    if warehouse_price_col is not None:
        label = warehouse_label_for_column(warehouses, warehouse_price_col)
        print(f"Using warehouse table: {label} (column {warehouse_price_col + 1})")

    rate_lookup = build_rate_lookup(
        rate_card_path,
        warehouse,
        warehouse_price_col=warehouse_price_col,
        interactive=interactive,
    )
    print(f"Loaded {len({k for k in rate_lookup if isinstance(k, str)})} rate card lookup keys")
    output_path = OUTPUT_DIR / f"{agreement_path.stem}_updated.xlsx"
    shutil.copy2(agreement_path, output_path)

    workbook = load_workbook(output_path)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )

    worksheet = workbook[sheet_name]
    updates, unmatched = apply_costs_to_service_rows(
        worksheet,
        rate_lookup,
        services,
        cost_mappings=cost_mappings,
    )
    workbook.save(output_path)
    workbook.close()

    return output_path, updates, unmatched


def main() -> None:
    agreement_file = choose_file(list_agreement_files(), "rate agreement")
    rate_card_file = choose_file(list_rate_card_files(), "rate card (processing)")

    output_path, updates, unmatched = apply_rate_card_to_agreement(
        agreement_file,
        rate_card_file,
        services=DEFAULT_SERVICES,
    )

    print(f"\nApplied rate card costs for services: {', '.join(DEFAULT_SERVICES)}")
    if updates:
        print("Updated cells:")
        for item in updates:
            source = item.get("match_source", "name")
            if item.get("action") == "highlighted_yellow":
                print(
                    f"  Row {item['row']}, {item['category']}: "
                    f"highlighted yellow (value kept: {item['old_value']})"
                )
            else:
                print(
                    f"  Row {item['row']}, {item['category']} [{source}]: "
                    f"{item['old_value']} -> {item['new_value']}"
                )
    else:
        print("No values were updated.")

    if unmatched:
        print("\nCosts not updated (no mapping/rate match):")
        seen_categories: set[str] = set()
        for item in unmatched:
            category = str(item["category"])
            if category in seen_categories:
                continue
            seen_categories.add(category)
            print(f"  {category}")

    print(f"\nSaved to: {output_path}")


if __name__ == "__main__":
    main()
